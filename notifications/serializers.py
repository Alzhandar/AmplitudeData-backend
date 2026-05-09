from __future__ import annotations

import io
from typing import Iterable, List

from rest_framework import serializers


class NotificationCitySerializer(serializers.Serializer):
    id = serializers.IntegerField()
    name_ru = serializers.CharField(allow_blank=True)
    name_kz = serializers.CharField(allow_blank=True)


class PushDispatchRequestSerializer(serializers.Serializer):
    target = serializers.ChoiceField(choices=("phones", "city"))
    phone_numbers = serializers.ListField(
        child=serializers.CharField(allow_blank=False),
        required=False,
        allow_empty=True,
    )
    city_id = serializers.IntegerField(required=False, min_value=1)
    excel_file = serializers.FileField(required=False, allow_null=True)

    title = serializers.CharField(max_length=255)
    body = serializers.CharField()
    title_kz = serializers.CharField(max_length=255, required=False, allow_blank=True, default="")
    body_kz = serializers.CharField(required=False, allow_blank=True, default="")
    notification_type = serializers.CharField(max_length=64, required=False, default="default")
    survey_id = serializers.IntegerField(required=False, allow_null=True)
    review_id = serializers.IntegerField(required=False, allow_null=True)

    def validate(self, attrs):
        target = attrs["target"]
        phone_numbers = attrs.get("phone_numbers") or []
        city_id = attrs.get("city_id")
        excel_file = attrs.get("excel_file")

        if target == "phones":
            normalized_numbers = _normalize_phone_numbers(phone_numbers)

            if excel_file is not None:
                filename = str(getattr(excel_file, "name", "")).lower()
                if filename and not filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
                    raise serializers.ValidationError({"excel_file": "Excel file must be .xlsx/.xlsm/.xltx/.xltm."})

                excel_numbers = _extract_phone_numbers_from_excel(excel_file.read())
                normalized_numbers = _merge_unique(normalized_numbers, excel_numbers)

            if not normalized_numbers:
                raise serializers.ValidationError({"phone_numbers": "Provide at least one valid phone number."})
            attrs["phone_numbers"] = normalized_numbers
            attrs["city_id"] = None

        if target == "city":
            if city_id is None:
                raise serializers.ValidationError({"city_id": "City is required."})
            attrs["phone_numbers"] = []
            attrs["excel_file"] = None

        # Excel file is a transport field used only for request parsing.
        attrs.pop("excel_file", None)

        return attrs


def _normalize_phone_numbers(values: Iterable[str]) -> List[str]:
    if isinstance(values, str):
        values = [values]

    normalized: List[str] = []
    seen = set()

    for index, raw in enumerate(values):
        phone = _normalize_phone(raw)
        if not phone:
            raise serializers.ValidationError(
                {
                    "phone_numbers": f"Invalid phone number at position {index + 1}: {raw}",
                }
            )

        if phone in seen:
            continue

        seen.add(phone)
        normalized.append(phone)

    return normalized


def _merge_unique(first: List[str], second: List[str]) -> List[str]:
    merged: List[str] = []
    seen = set()
    for value in [*first, *second]:
        if value in seen:
            continue
        seen.add(value)
        merged.append(value)
    return merged


def _extract_phone_numbers_from_excel(binary: bytes) -> List[str]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise serializers.ValidationError({"excel_file": "openpyxl library is not available."}) from exc

    try:
        workbook = load_workbook(filename=io.BytesIO(binary), data_only=True)
    except Exception as exc:
        raise serializers.ValidationError({"excel_file": "Failed to read Excel file."}) from exc

    sheet = workbook.active
    normalized: List[str] = []
    seen = set()

    for row_index, row in enumerate(sheet.iter_rows(min_row=1, min_col=1, max_col=1), start=1):
        value = row[0].value
        if value is None:
            continue

        raw = str(value).strip()
        if not raw or _is_phone_header_label(raw):
            continue

        phone = _normalize_phone(raw)
        if not phone:
            raise serializers.ValidationError({"excel_file": f"Invalid phone number at row {row_index}: {raw}"})

        if phone in seen:
            continue

        seen.add(phone)
        normalized.append(phone)

    return normalized


def _normalize_phone(value: str) -> str:
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    if not digits:
        return ""

    if len(digits) == 11 and digits.startswith("8"):
        digits = f"7{digits[1:]}"

    if len(digits) != 11 or not digits.startswith("7"):
        return ""

    return digits


def _is_phone_header_label(value: str) -> bool:
    cleaned = "".join(ch.lower() for ch in str(value) if ch.isalnum() or ch in {" ", "_", "-"})
    cleaned = cleaned.strip().replace("-", " ").replace("_", " ")
    collapsed = " ".join(cleaned.split())
    compact = collapsed.replace(" ", "")

    candidates = {
        "phone",
        "phones",
        "phone number",
        "phone numbers",
        "phonenumber",
        "phonenumbers",
        "телефон",
        "телефоны",
        "номер телефона",
        "номертелефона",
    }
    return collapsed in candidates or compact in candidates
