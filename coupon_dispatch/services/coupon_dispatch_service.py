from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from datetime import date
from typing import Any, Dict, List, Optional, Tuple

from django.utils import timezone

from coupon_dispatch.models import (
    CouponDispatchInputSource,
    CouponDispatchJob,
    CouponDispatchMode,
    CouponDispatchJobResult,
    CouponDispatchJobStatus,
)
from utils.avatariya_client import AvatariyaClient
from utils.mobile_client import MobileClient

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class MarketingSaleOption:
    id: int
    name: str
    available_coupons: int


@dataclass(frozen=True)
class ParsedPhoneRow:
    phone_raw: str
    phone_normalized: str
    valid: bool
    error_message: str


@dataclass(frozen=True)
class ParsedPhoneCouponRow:
    phone_raw: str
    phone_normalized: str
    coupon_code: str
    valid: bool
    error_message: str


class CouponDispatchService:
    def __init__(
        self,
        avatariya_client: Optional[AvatariyaClient] = None,
        mobile_client: Optional[MobileClient] = None,
    ) -> None:
        self.avatariya_client = avatariya_client or AvatariyaClient()
        self.mobile_client = mobile_client or MobileClient()

    def list_marketing_sales_with_available_coupons(self, search: str = '') -> List[MarketingSaleOption]:
        raw_items = self.avatariya_client.list_coupon_assign_marketing_sales()
        normalized_search = str(search or '').strip().lower()
        options: List[MarketingSaleOption] = []

        for item in raw_items:
            sale_id = self._to_int(item.get('id'))
            if sale_id is None:
                continue

            status = item.get('status')
            if status is False:
                continue

            sale_name = str(item.get('name') or '').strip()
            if normalized_search and normalized_search not in sale_name.lower():
                continue

            free_coupons_count = self._to_int(item.get('free_coupons_count')) or 0
            options.append(
                MarketingSaleOption(
                    id=sale_id,
                    name=sale_name,
                    available_coupons=max(free_coupons_count, 0),
                )
            )

        options.sort(key=lambda value: (-value.available_coupons, value.name.lower()))
        return options

    def create_job(
        self,
        *,
        user,
        title: str,
        dispatch_mode: str,
        marketing_sale_id: Optional[int],
        marketing_sale_name: str,
        valid_until: date,
        phones_text: str,
        source_file,
    ) -> CouponDispatchJob:
        normalized_title = (title or '').strip()
        if not normalized_title:
            raise ValueError('Coupon title is required')

        normalized_mode = str(dispatch_mode or '').strip()
        has_text = bool((phones_text or '').strip())
        has_file = bool(source_file)

        if normalized_mode not in {'marketing_sale', 'predefined_coupon'}:
            raise ValueError('dispatch_mode must be marketing_sale or predefined_coupon')

        if normalized_mode == 'marketing_sale':
            if not marketing_sale_id or int(marketing_sale_id) <= 0:
                raise ValueError('marketing_sale_id is required for marketing_sale mode')
            if not has_text and not has_file:
                raise ValueError('Provide manual phone numbers or an Excel file')
            normalized_marketing_sale_id = int(marketing_sale_id)
            normalized_marketing_sale_name = (marketing_sale_name or '').strip()
            normalized_source_text = phones_text or ''
        else:
            if not has_file:
                raise ValueError('Excel file is required for predefined_coupon mode')
            normalized_marketing_sale_id = 0
            normalized_marketing_sale_name = ''
            normalized_source_text = ''

        if has_text and has_file:
            input_source = CouponDispatchInputSource.MIXED
        elif has_file:
            input_source = CouponDispatchInputSource.EXCEL
        else:
            input_source = CouponDispatchInputSource.MANUAL

        return CouponDispatchJob.objects.create(
            title=normalized_title,
            dispatch_mode=normalized_mode,
            marketing_sale_id=normalized_marketing_sale_id,
            marketing_sale_name=normalized_marketing_sale_name,
            valid_until=valid_until,
            initiated_by=user,
            input_source=input_source,
            source_text=normalized_source_text,
            source_file=source_file,
            status=CouponDispatchJobStatus.PENDING,
        )

    def process_job(self, job_id: int) -> Dict[str, Any]:
        job = CouponDispatchJob.objects.filter(id=job_id).first()
        if not job:
            raise ValueError(f'Coupon dispatch job {job_id} does not exist')

        now = timezone.now()
        claimed = CouponDispatchJob.objects.filter(
            id=job_id,
            status=CouponDispatchJobStatus.PENDING,
            coupons_assigned=0,
        ).update(
            status=CouponDispatchJobStatus.PROCESSING,
            started_at=now,
            finished_at=None,
            error_log='',
            updated_at=now,
        )

        if claimed == 0:
            job = CouponDispatchJob.objects.filter(id=job_id).first()
            logger.info(
                'job_skipped',
                extra={
                    'job_id': job_id,
                    'status': job.status if job else 'missing',
                    'coupons_assigned': job.coupons_assigned if job else 'n/a',
                },
            )
            return {
                'status': 'skipped',
                'reason': 'non_pending_or_already_assigned',
                'job_status': job.status if job else 'missing',
            }

        job = CouponDispatchJob.objects.filter(id=job_id).first()
        logger.info('job_started', extra={'job_id': job_id, 'mode': job.dispatch_mode if job else 'unknown'})

        try:
            mode = str(job.dispatch_mode or '').strip()
            if mode == CouponDispatchMode.PREDEFINED_COUPON:
                result = self._process_predefined_coupon_job(job)
            elif mode == CouponDispatchMode.MARKETING_SALE and job.marketing_sale_id > 0:
                result = self._process_marketing_sale_job(job)
            elif mode == CouponDispatchMode.MARKETING_SALE and job.marketing_sale_id <= 0 and job.source_file:
                logger.warning(
                    'job_mode_fallback',
                    extra={'job_id': job.id, 'reason': 'marketing_sale_mode_with_empty_id'},
                )
                result = self._process_predefined_coupon_job(job)
            else:
                # Backward-compatible fallback for historical rows without dispatch_mode.
                result = self._process_marketing_sale_job(job) if job.marketing_sale_id > 0 else self._process_predefined_coupon_job(job)

            logger.info(
                'job_finished',
                extra={
                    'job_id': job.id,
                    'status': result['status'],
                    'coupons_assigned': result['coupons_assigned'],
                    'errors_count': result['errors_count'],
                },
            )
            return result
        except Exception as exc:
            logger.exception('job_failed', extra={'job_id': job.id})
            job.status = CouponDispatchJobStatus.FAILED
            job.finished_at = timezone.now()
            job.error_log = f'{job.error_log}\nunhandled_exception'.strip()
            job.save(update_fields=['status', 'finished_at', 'error_log', 'updated_at'])
            raise

    def _process_marketing_sale_job(self, job: CouponDispatchJob) -> Dict[str, Any]:
        raw_phones = self._collect_raw_phones(job)
        parsed_rows = [self._parse_phone(phone) for phone in raw_phones]

        valid_rows: List[ParsedPhoneRow] = []
        seen_phones = set()
        result_rows: List[CouponDispatchJobResult] = []

        for parsed in parsed_rows:
            if not parsed.valid:
                result_rows.append(
                    CouponDispatchJobResult(
                        job=job,
                        phone_raw=parsed.phone_raw,
                        phone_normalized='',
                        success=False,
                        error_message=parsed.error_message,
                    )
                )
                continue

            if parsed.phone_normalized in seen_phones:
                result_rows.append(
                    CouponDispatchJobResult(
                        job=job,
                        phone_raw=parsed.phone_raw,
                        phone_normalized=parsed.phone_normalized,
                        success=False,
                        error_message='duplicate_phone',
                    )
                )
                continue

            seen_phones.add(parsed.phone_normalized)
            valid_rows.append(parsed)

        pre_validation_errors = sum(1 for row in result_rows if not row.success)
        job.total_phones = len(raw_phones)
        job.unique_phones = len(valid_rows)
        job.errors_count = pre_validation_errors
        job.save(update_fields=['total_phones', 'unique_phones', 'errors_count', 'updated_at'])

        guests_found = 0
        coupons_assigned = 0
        mobile_sent_count = 0
        assign_responses: List[Dict[str, Any]] = []

        for row in valid_rows:
            try:
                assign_result = self.avatariya_client.assign_coupon_via_admin(
                    marketing_sale_id=job.marketing_sale_id,
                    phone_number=row.phone_normalized,
                    amount=job.title,
                    valid_until=job.valid_until.isoformat() if job.valid_until else '',
                )
            except Exception as exc:
                logger.warning(
                    'assign_api_error',
                    extra={'job_id': job.id, 'phone': row.phone_normalized, 'detail': str(exc)},
                )
                result_rows.append(
                    CouponDispatchJobResult(
                        job=job,
                        phone_raw=row.phone_raw,
                        phone_normalized=row.phone_normalized,
                        success=False,
                        error_message='assign_api_error',
                    )
                )
                continue

            assigned = bool(assign_result.get('assigned'))
            guest_id = self._to_int(assign_result.get('guest_id'))
            coupon_id = self._to_int(assign_result.get('coupon_id'))
            coupon_code = str(assign_result.get('coupon_code') or '').strip()
            mobile_sent = bool(assign_result.get('mobile_sent'))

            if guest_id is not None:
                guests_found += 1
            if mobile_sent:
                mobile_sent_count += 1

            if assigned:
                coupons_assigned += 1
                result_rows.append(
                    CouponDispatchJobResult(
                        job=job,
                        phone_raw=row.phone_raw,
                        phone_normalized=row.phone_normalized,
                        guest_id=guest_id,
                        coupon_id=coupon_id,
                        coupon_code=coupon_code,
                        success=True,
                        error_message='',
                    )
                )
            else:
                result_rows.append(
                    CouponDispatchJobResult(
                        job=job,
                        phone_raw=row.phone_raw,
                        phone_normalized=row.phone_normalized,
                        guest_id=guest_id,
                        coupon_id=coupon_id,
                        coupon_code=coupon_code,
                        success=False,
                        error_message='coupon_not_assigned',
                    )
                )

            assign_responses.append(
                {
                    'phone': row.phone_normalized,
                    'assigned': assigned,
                    'message': str(assign_result.get('message') or ''),
                    'mobile_message': str(assign_result.get('mobile_message') or ''),
                    'mobile_sent': mobile_sent,
                    'coupon_id': coupon_id,
                    'coupon_code': coupon_code,
                }
            )

        free_count_after = self._get_free_coupons_count_for_sale(job.marketing_sale_id)
        return self._finalize_job(
            job=job,
            result_rows=result_rows,
            total_phones=len(raw_phones),
            unique_phones=len(valid_rows),
            guests_found=guests_found,
            available_coupons=free_count_after,
            coupons_assigned=coupons_assigned,
            mobile_sent_count=mobile_sent_count,
            api_path='/api/v1/admin/coupon-assign/assign/',
            responses=assign_responses,
        )

    def _process_predefined_coupon_job(self, job: CouponDispatchJob) -> Dict[str, Any]:
        if not job.source_file:
            raise ValueError('Excel file is required for predefined_coupon mode')

        raw_rows = self._read_excel_phone_coupon_rows(job.source_file.read())
        parsed_rows = [self._parse_phone_coupon_row(phone_raw, coupon_raw) for phone_raw, coupon_raw in raw_rows]

        valid_rows: List[ParsedPhoneCouponRow] = []
        seen_phones = set()
        result_rows: List[CouponDispatchJobResult] = []

        for parsed in parsed_rows:
            if not parsed.valid:
                result_rows.append(
                    CouponDispatchJobResult(
                        job=job,
                        phone_raw=parsed.phone_raw,
                        phone_normalized=parsed.phone_normalized,
                        coupon_code=parsed.coupon_code,
                        success=False,
                        error_message=parsed.error_message,
                    )
                )
                continue

            if parsed.phone_normalized in seen_phones:
                result_rows.append(
                    CouponDispatchJobResult(
                        job=job,
                        phone_raw=parsed.phone_raw,
                        phone_normalized=parsed.phone_normalized,
                        coupon_code=parsed.coupon_code,
                        success=False,
                        error_message='duplicate_phone',
                    )
                )
                continue

            seen_phones.add(parsed.phone_normalized)
            valid_rows.append(parsed)

        pre_validation_errors = sum(1 for row in result_rows if not row.success)
        job.total_phones = len(raw_rows)
        job.unique_phones = len(valid_rows)
        job.errors_count = pre_validation_errors
        job.save(update_fields=['total_phones', 'unique_phones', 'errors_count', 'updated_at'])

        coupons_assigned = 0
        mobile_sent_count = 0
        api_responses: List[Dict[str, Any]] = []

        for row in valid_rows:
            try:
                api_result = self.mobile_client.create_order_coupon_info_bulk_item(
                    coupon=row.coupon_code,
                    phone_number=row.phone_normalized,
                    amount=job.title,
                    valid_until=job.valid_until.isoformat() if job.valid_until else '',
                    is_mobile=False,
                )
            except Exception as exc:
                logger.warning(
                    'mobile_api_error',
                    extra={'job_id': job.id, 'phone': row.phone_normalized, 'coupon': row.coupon_code, 'detail': str(exc)},
                )
                result_rows.append(
                    CouponDispatchJobResult(
                        job=job,
                        phone_raw=row.phone_raw,
                        phone_normalized=row.phone_normalized,
                        coupon_code=row.coupon_code,
                        success=False,
                        error_message='mobile_api_error',
                    )
                )
                continue

            coupons_assigned += 1
            mobile_sent_count += 1
            result_rows.append(
                CouponDispatchJobResult(
                    job=job,
                    phone_raw=row.phone_raw,
                    phone_normalized=row.phone_normalized,
                    coupon_code=row.coupon_code,
                    success=True,
                    error_message='',
                )
            )
            api_responses.append(
                {
                    'phone': row.phone_normalized,
                    'coupon_code': row.coupon_code,
                    'response': api_result,
                }
            )

        return self._finalize_job(
            job=job,
            result_rows=result_rows,
            total_phones=len(raw_rows),
            unique_phones=len(valid_rows),
            guests_found=0,
            available_coupons=0,
            coupons_assigned=coupons_assigned,
            mobile_sent_count=mobile_sent_count,
            api_path='/api/prizes/order-coupon-info/bulk-create/',
            responses=api_responses,
        )

    def _finalize_job(
        self,
        *,
        job: CouponDispatchJob,
        result_rows: List[CouponDispatchJobResult],
        total_phones: int,
        unique_phones: int,
        guests_found: int,
        available_coupons: int,
        coupons_assigned: int,
        mobile_sent_count: int,
        api_path: str,
        responses: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        CouponDispatchJobResult.objects.filter(job=job).delete()
        if result_rows:
            CouponDispatchJobResult.objects.bulk_create(result_rows, batch_size=500)

        errors_count = sum(1 for row in result_rows if not row.success)
        error_log_items = [row.error_message for row in result_rows if row.error_message]
        mobile_api_sent = mobile_sent_count > 0
        mobile_api_response: Dict[str, Any] = {
            'assign_api_path': api_path,
            'attempted': unique_phones,
            'mobile_sent_count': mobile_sent_count,
            'valid_until': job.valid_until.isoformat() if job.valid_until else None,
            'responses_sample': responses[:100],
        }

        job.status = CouponDispatchJobStatus.COMPLETED
        job.total_phones = total_phones
        job.unique_phones = unique_phones
        job.guests_found = guests_found
        job.available_coupons = available_coupons
        job.coupons_assigned = coupons_assigned
        job.errors_count = errors_count
        job.mobile_api_sent = mobile_api_sent
        job.mobile_api_response = mobile_api_response
        job.mobile_api_sent_at = timezone.now() if mobile_api_sent else None
        job.error_log = '\n'.join(error_log_items)[:8000]
        job.finished_at = timezone.now()
        job.save(
            update_fields=[
                'status',
                'total_phones',
                'unique_phones',
                'guests_found',
                'available_coupons',
                'coupons_assigned',
                'errors_count',
                'mobile_api_sent',
                'mobile_api_response',
                'mobile_api_sent_at',
                'error_log',
                'finished_at',
                'updated_at',
            ]
        )

        return {
            'job_id': job.id,
            'status': job.status,
            'coupons_assigned': job.coupons_assigned,
            'errors_count': job.errors_count,
        }

    def _get_free_coupons_count_for_sale(self, marketing_sale_id: int) -> int:
        if not marketing_sale_id or marketing_sale_id <= 0:
            return 0

        try:
            sales = self.avatariya_client.list_coupon_assign_marketing_sales()
        except Exception:
            logger.exception('Failed to fetch marketing sales for free_coupons_count')
            return 0

        for item in sales:
            sale_id = self._to_int(item.get('id'))
            if sale_id == marketing_sale_id:
                return self._to_int(item.get('free_coupons_count')) or 0
        return 0

    def _collect_raw_phones(self, job: CouponDispatchJob) -> List[str]:
        values: List[str] = []

        if job.source_text.strip():
            values.extend(self._split_text_phones(job.source_text))

        if job.source_file:
            values.extend(self._read_excel_phones(job.source_file.read()))

        return [value for value in values if value]

    def _split_text_phones(self, source_text: str) -> List[str]:
        text = source_text.replace('\r', '\n').replace(';', '\n').replace(',', '\n')
        values: List[str] = []
        for raw_line in text.split('\n'):
            line = raw_line.strip()
            if not line:
                continue
            if self._is_phone_header_label(line):
                continue
            values.append(line)
        return values

    def _read_excel_phones(self, binary: bytes) -> List[str]:
        from openpyxl import load_workbook

        workbook = load_workbook(filename=io.BytesIO(binary), data_only=True)
        sheet = workbook.active
        phones: List[str] = []
        for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1):
            value = row[0].value
            if value is None:
                continue
            normalized = str(value).strip()
            if normalized:
                if self._is_phone_header_label(normalized):
                    continue
                phones.append(normalized)
        return phones

    def _read_excel_phone_coupon_rows(self, binary: bytes) -> List[Tuple[str, str]]:
        from openpyxl import load_workbook

        workbook = load_workbook(filename=io.BytesIO(binary), data_only=True)
        sheet = workbook.active
        rows: List[Tuple[str, str]] = []

        for row in sheet.iter_rows(min_row=1, min_col=1, max_col=2):
            raw_phone = str(row[0].value or '').strip()
            raw_coupon = str(row[1].value or '').strip() if len(row) > 1 else ''

            if not raw_phone and not raw_coupon:
                continue

            if self._is_phone_header_label(raw_phone) and self._is_coupon_header_label(raw_coupon):
                continue

            rows.append((raw_phone, raw_coupon))

        return rows

    def _is_phone_header_label(self, value: str) -> bool:
        cleaned = ''.join(ch.lower() for ch in str(value) if ch.isalnum() or ch in {' ', '_', '-'})
        cleaned = cleaned.strip().replace('-', ' ').replace('_', ' ')
        collapsed = ' '.join(cleaned.split())
        compact = collapsed.replace(' ', '')

        candidates = {
            'phone',
            'phones',
            'phone number',
            'phone numbers',
            'phonenumber',
            'phonenumbers',
            'телефон',
            'телефоны',
            'номер телефона',
            'номертелефона',
        }
        return collapsed in candidates or compact in candidates

    def _is_coupon_header_label(self, value: str) -> bool:
        cleaned = ''.join(ch.lower() for ch in str(value) if ch.isalnum() or ch in {' ', '_', '-'})
        cleaned = cleaned.strip().replace('-', ' ').replace('_', ' ')
        collapsed = ' '.join(cleaned.split())
        compact = collapsed.replace(' ', '')

        candidates = {
            'coupon',
            'coupon code',
            'couponcode',
            'code',
            'promo code',
            'promocode',
            'купон',
            'код',
            'код купона',
            'кодкупона',
        }
        return collapsed in candidates or compact in candidates

    def _parse_phone(self, phone_raw: str) -> ParsedPhoneRow:
        normalized = ''.join(ch for ch in str(phone_raw or '') if ch.isdigit())
        if not normalized:
            return ParsedPhoneRow(phone_raw=phone_raw, phone_normalized='', valid=False, error_message='phone_empty')

        if len(normalized) == 11 and normalized.startswith('8'):
            normalized = f'7{normalized[1:]}'

        if len(normalized) != 11 or not normalized.startswith('7'):
            return ParsedPhoneRow(
                phone_raw=phone_raw,
                phone_normalized='',
                valid=False,
                error_message='invalid_phone_format',
            )

        return ParsedPhoneRow(phone_raw=phone_raw, phone_normalized=normalized, valid=True, error_message='')

    def _parse_phone_coupon_row(self, phone_raw: str, coupon_raw: str) -> ParsedPhoneCouponRow:
        parsed_phone = self._parse_phone(phone_raw)
        if not parsed_phone.valid:
            return ParsedPhoneCouponRow(
                phone_raw=phone_raw,
                phone_normalized=parsed_phone.phone_normalized,
                coupon_code=str(coupon_raw or '').strip(),
                valid=False,
                error_message=parsed_phone.error_message,
            )

        coupon_code = str(coupon_raw or '').strip()
        if not coupon_code:
            return ParsedPhoneCouponRow(
                phone_raw=phone_raw,
                phone_normalized=parsed_phone.phone_normalized,
                coupon_code='',
                valid=False,
                error_message='coupon_code_empty',
            )

        return ParsedPhoneCouponRow(
            phone_raw=phone_raw,
            phone_normalized=parsed_phone.phone_normalized,
            coupon_code=coupon_code,
            valid=True,
            error_message='',
        )

    def _to_int(self, value: Any) -> Optional[int]:
        try:
            if value is None:
                return None
            return int(value)
        except (TypeError, ValueError):
            return None
