from __future__ import annotations

import io
import re
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Dict, List, Optional, Sequence

from django.db import transaction
from django.db.utils import OperationalError, ProgrammingError
from django.utils import timezone as dj_timezone
from openpyxl import load_workbook

from bonus_transactions.models import (
    BonusTransactionJob,
    BonusTransactionJobResult,
    BonusTransactionSettings,
    BonusTransactionJobStatus,
    BonusTransactionInputSource,
)
from utils.avatariya_client import AvatariyaClient


@dataclass
class _PhoneRow:
    phone_raw: str
    phone_normalized: str


class BonusTransactionService:
    def __init__(self, avatariya_client: Optional[AvatariyaClient] = None):
        self.avatariya = avatariya_client or AvatariyaClient()

    @transaction.atomic
    def create_job(
        self,
        *,
        initiated_by,
        description: str,
        amount: int,
        start_date,
        expiration_date,
        phones_text: str,
        excel_file,
    ) -> BonusTransactionJob:
        has_text = bool(str(phones_text or '').strip())
        has_file = excel_file is not None

        if has_text and has_file:
            input_source = BonusTransactionInputSource.MIXED
        elif has_file:
            input_source = BonusTransactionInputSource.EXCEL
        else:
            input_source = BonusTransactionInputSource.MANUAL

        job = BonusTransactionJob.objects.create(
            initiated_by=initiated_by,
            description=description.strip(),
            amount=amount,
            start_date=start_date,
            expiration_date=expiration_date,
            base_id_prefix=self._get_base_id_prefix(),
            source_text=str(phones_text or ''),
            source_file=excel_file,
            input_source=input_source,
            status=BonusTransactionJobStatus.PENDING,
        )
        return job

    def process_job(self, job_id: int) -> BonusTransactionJob:
        job = BonusTransactionJob.objects.get(pk=job_id)

        # Safety guard: process only pending jobs to avoid accidental duplicate accruals.
        if job.status != BonusTransactionJobStatus.PENDING:
            return job

        # Additional guard for any inconsistent status transitions.
        if job.cashbacks_created > 0:
            return job

        now = dj_timezone.now()
        claimed = BonusTransactionJob.objects.filter(
            pk=job_id,
            status=BonusTransactionJobStatus.PENDING,
            cashbacks_created=0,
        ).update(
            status=BonusTransactionJobStatus.PROCESSING,
            started_at=now,
            finished_at=None,
            error_log='',
            external_api_response={},
            total_phones=0,
            unique_phones=0,
            guests_found=0,
            cashbacks_created=0,
            errors_count=0,
            updated_at=now,
        )
        if claimed == 0:
            return BonusTransactionJob.objects.get(pk=job_id)

        job = BonusTransactionJob.objects.get(pk=job_id)

        try:
            raw_phones = self._collect_phone_candidates(job)
            parsed_rows, parse_errors = self._parse_and_dedupe(raw_phones)

            guest_found_count = 0
            success_count = 0
            result_rows: List[BonusTransactionJobResult] = []
            api_outcomes: List[Dict] = []

            for message in parse_errors:
                result_rows.append(
                    BonusTransactionJobResult(
                        job=job,
                        phone_raw='',
                        phone_normalized='',
                        success=False,
                        error_message=message,
                    )
                )

            for index, row in enumerate(parsed_rows, start=1):
                try:
                    guest_payload = self.avatariya.find_guest_by_phone(row.phone_normalized)
                    guest_id = self._extract_guest_id(guest_payload)
                    if guest_id is None:
                        result_rows.append(
                            BonusTransactionJobResult(
                                job=job,
                                phone_raw=row.phone_raw,
                                phone_normalized=row.phone_normalized,
                                success=False,
                                error_message='Guest not found by phone',
                            )
                        )
                        continue

                    guest_found_count += 1
                    doc_guid = str(uuid.uuid4())
                    base_id = self._build_base_id(job.base_id_prefix, guest_id, index)
                    payload = {
                        'doc_guid': doc_guid,
                        'guest': int(guest_id),
                        'park_id': None,
                        'amount': int(job.amount),
                        'transaction_date': self._utc_iso_now(),
                        'date': self._utc_iso_now(),
                        'base_id': base_id,
                        'type': int(job.bonus_type),
                        'registration_bonus': bool(job.registration_bonus),
                        'start_date': job.start_date.isoformat(),
                        'expiration_date': job.expiration_date.isoformat(),
                        'description': job.description,
                    }

                    api_response = self.avatariya.create_cashback(payload)
                    success_count += 1
                    api_outcomes.append({'phone': row.phone_normalized, 'ok': True})
                    result_rows.append(
                        BonusTransactionJobResult(
                            job=job,
                            phone_raw=row.phone_raw,
                            phone_normalized=row.phone_normalized,
                            guest_id=guest_id,
                            doc_guid=doc_guid,
                            base_id=base_id,
                            success=True,
                            cashback_payload=payload,
                            cashback_response=api_response if isinstance(api_response, dict) else {'raw': api_response},
                        )
                    )
                except Exception as exc:
                    api_outcomes.append({'phone': row.phone_normalized, 'ok': False, 'error': str(exc)})
                    result_rows.append(
                        BonusTransactionJobResult(
                            job=job,
                            phone_raw=row.phone_raw,
                            phone_normalized=row.phone_normalized,
                            success=False,
                            error_message=str(exc),
                        )
                    )

            BonusTransactionJobResult.objects.filter(job=job).delete()
            if result_rows:
                BonusTransactionJobResult.objects.bulk_create(result_rows, batch_size=300)

            total_errors = sum(1 for item in result_rows if not item.success)
            error_messages = [item.error_message for item in result_rows if item.error_message]
            job.total_phones = len(raw_phones)
            job.unique_phones = len(parsed_rows)
            job.guests_found = guest_found_count
            job.cashbacks_created = success_count
            job.errors_count = total_errors
            job.status = BonusTransactionJobStatus.COMPLETED
            job.finished_at = dj_timezone.now()
            job.error_log = '\n'.join(error_messages[:500])
            job.external_api_response = {
                'processed': len(parsed_rows),
                'success': success_count,
                'errors': total_errors,
                'sample': api_outcomes[:100],
            }
            job.save(
                update_fields=[
                    'total_phones',
                    'unique_phones',
                    'guests_found',
                    'cashbacks_created',
                    'errors_count',
                    'status',
                    'finished_at',
                    'error_log',
                    'external_api_response',
                    'updated_at',
                ]
            )
            return job
        except Exception as exc:
            job.status = BonusTransactionJobStatus.FAILED
            job.finished_at = dj_timezone.now()
            job.error_log = str(exc)
            job.errors_count = max(1, job.errors_count)
            job.save(update_fields=['status', 'finished_at', 'error_log', 'errors_count', 'updated_at'])
            raise

    def _collect_phone_candidates(self, job: BonusTransactionJob) -> List[str]:
        candidates: List[str] = []

        if job.source_text:
            candidates.extend(self._extract_phones_from_text(job.source_text))

        if job.source_file:
            candidates.extend(self._extract_phones_from_excel(job.source_file.read()))

        return [value for value in candidates if str(value).strip()]

    def _parse_and_dedupe(self, values: Sequence[str]) -> tuple[List[_PhoneRow], List[str]]:
        parsed_rows: List[_PhoneRow] = []
        errors: List[str] = []
        seen: set[str] = set()

        for raw in values:
            normalized = self._normalize_phone(raw)
            if not normalized:
                errors.append(f'Invalid phone number: {raw}')
                continue
            if normalized in seen:
                continue
            seen.add(normalized)
            parsed_rows.append(_PhoneRow(phone_raw=str(raw).strip(), phone_normalized=normalized))

        return parsed_rows, errors

    def _extract_phones_from_text(self, text: str) -> List[str]:
        separators = r'[\n,;\t\r ]+'
        parts = re.split(separators, text)
        return [part.strip() for part in parts if part and part.strip()]

    def _extract_phones_from_excel(self, file_bytes: bytes) -> List[str]:
        if not file_bytes:
            return []

        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        phones: List[str] = []

        for sheet in workbook.worksheets:
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                row_values = [self._cell_to_str(cell) for cell in row]
                for value in row_values:
                    if not value:
                        continue
                    if row_idx == 1 and self._is_phone_header(value):
                        continue
                    phones.append(value)
                    break

        return phones

    def _normalize_phone(self, value: str) -> str:
        if value is None:
            return ''

        digits = ''.join(ch for ch in str(value) if ch.isdigit())
        if len(digits) == 11 and digits.startswith('8'):
            return f'7{digits[1:]}'
        if len(digits) == 10 and digits.startswith('0'):
            return ''
        if len(digits) == 10:
            return f'7{digits}'
        if len(digits) == 11 and digits.startswith('7'):
            return digits
        return ''

    def _is_phone_header(self, value: str) -> bool:
        normalized = str(value).strip().lower()
        labels = {
            'phone',
            'phones',
            'phone number',
            'phone_number',
            'телефон',
            'номер',
            'номер телефона',
            'mobile',
            'msisdn',
        }
        return normalized in labels

    def _cell_to_str(self, value) -> str:
        if value is None:
            return ''
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _extract_guest_id(self, payload) -> Optional[int]:
        if isinstance(payload, dict):
            raw_id = payload.get('id')
            if raw_id is None and isinstance(payload.get('data'), dict):
                raw_id = payload['data'].get('id')
            if raw_id is None and isinstance(payload.get('result'), dict):
                raw_id = payload['result'].get('id')
            if raw_id is None and isinstance(payload.get('results'), list) and payload['results']:
                first = payload['results'][0]
                if isinstance(first, dict):
                    raw_id = first.get('id')
            try:
                return int(raw_id)
            except (TypeError, ValueError):
                return None
        return None

    def _build_base_id(self, prefix: str, guest_id: int, index: int) -> str:
        normalized_prefix = str(prefix or '').strip() or 'bonus'
        composed = f'{normalized_prefix}-{guest_id}-{index}'
        return composed[:255]

    def _get_base_id_prefix(self) -> str:
        try:
            settings_obj, _ = BonusTransactionSettings.objects.get_or_create(singleton_guard=True)
            return str(settings_obj.base_id_prefix or '').strip() or 'bonus'
        except (ProgrammingError, OperationalError):
            return 'bonus'

    def _utc_iso_now(self) -> str:
        now = datetime.now(timezone.utc)
        return now.isoformat(timespec='milliseconds').replace('+00:00', 'Z')
